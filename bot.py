"""
🍽️ بوت وصفات الطعام العالمي
- يبحث في يوتيب عن الوصفات
- يولّد وصفة تفصيلية جميلة بالمكونات وطريقة التحضير عبر Claude AI
- يحفظ كل شيء في Excel
- جاهز للرفع على Render
"""

import os, json, logging, re, time, threading
from datetime import datetime
from http.server import HTTPServer, BaseHTTPRequestHandler

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import requests

# ══════════════════════════════════════════════════════
#  إعدادات
# ══════════════════════════════════════════════════════
BOT_TOKEN     = os.environ.get("TELEGRAM_TOKEN",  "8719774473:AAG6_COb6UElTsmzxlJJNaltmrJoL5QsqvQ")
ANTHROPIC_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
EXCEL_FILE    = "wosafat.xlsx"
PORT          = int(os.environ.get("PORT", 8080))

logging.basicConfig(format="%(asctime)s | %(levelname)s | %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)
API = f"https://api.telegram.org/bot{BOT_TOKEN}"

# ══════════════════════════════════════════════════════
#  بيانات ثابتة
# ══════════════════════════════════════════════════════
MEAL_TYPES = [
    "وجبة فطور", "وجبة غداء", "وجبة عشاء",
    "حلويات وتحلية", "مشروبات وعصائر", "شوربة وحساء",
    "سلطات", "مقبلات وسناكس", "خبز وفطائر",
    "أطباق بحرية", "لحوم ودجاج", "أكل نباتي",
]

WORLD_CUISINES = [
    "عربي", "مصري", "سعودي", "مغربي", "لبناني", "سوري",
    "تركي", "إيطالي", "فرنسي", "هندي", "صيني", "ياباني",
    "كوري", "تايلاندي", "مكسيكي", "أمريكي", "يوناني",
    "إسباني", "بريطاني", "إثيوبي", "إيراني", "باكستاني",
    "روسي", "برازيلي", "أرجنتيني", "عالمي",
]

CONTINENT_MAP = {
    "مصري":"أفريقيا", "مغربي":"أفريقيا", "إثيوبي":"أفريقيا",
    "سعودي":"آسيا", "عربي":"آسيا", "لبناني":"آسيا", "سوري":"آسيا",
    "تركي":"آسيا", "هندي":"آسيا", "صيني":"آسيا", "ياباني":"آسيا",
    "كوري":"آسيا", "تايلاندي":"آسيا", "إيراني":"آسيا", "باكستاني":"آسيا",
    "إيطالي":"أوروبا", "فرنسي":"أوروبا", "يوناني":"أوروبا",
    "إسباني":"أوروبا", "بريطاني":"أوروبا", "روسي":"أوروبا",
    "مكسيكي":"أمريكا اللاتينية", "برازيلي":"أمريكا اللاتينية", "أرجنتيني":"أمريكا اللاتينية",
    "أمريكي":"أمريكا الشمالية",
    "عالمي":"العالم",
}

user_states  = {}
search_cache = {}
recipe_cache = {}

# ══════════════════════════════════════════════════════
#  Excel
# ══════════════════════════════════════════════════════
COLORS = {
    "title_bg":  "1A1A2E",
    "header_bg": "16213E",
    "row_odd":   "F3F0FF",
    "row_even":  "FFFFFF",
    "link":      "1565C0",
    "border":    "C9C0E8",
}

HEADERS = [
    "رقم", "اسم الوصفة", "النوع", "الدولة", "القارة",
    "المكونات", "طريقة التحضير", "رابط يوتيب",
    "القناة", "المدة (د)", "المشاهدات", "تاريخ الإضافة", "ملاحظات",
]

def _border():
    s = Side(style="thin", color=COLORS["border"])
    return Border(left=s, right=s, top=s, bottom=s)

def _init_sheet(ws):
    ws.sheet_view.rightToLeft = True
    ws.merge_cells("A1:M1")
    c = ws["A1"]
    c.value = "قاعدة بيانات وصفات الطعام العالمية"
    c.font = Font(bold=True, size=16, color="FFFFFF", name="Arial")
    c.fill = PatternFill("solid", start_color=COLORS["title_bg"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 42
    b = _border()
    for i, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=i)
        cell.value = h
        cell.font = Font(bold=True, size=10, color="FFFFFF", name="Arial")
        cell.fill = PatternFill("solid", start_color=COLORS["header_bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = b
    ws.row_dimensions[2].height = 28
    for i, w in enumerate([5, 24, 14, 16, 14, 40, 55, 38, 20, 8, 12, 16, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"

def get_or_create_wb():
    if os.path.exists(EXCEL_FILE):
        return openpyxl.load_workbook(EXCEL_FILE)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "الوصفات"
    _init_sheet(ws)
    wb.save(EXCEL_FILE)
    return wb

def save_recipe(r: dict) -> int:
    wb = get_or_create_wb()
    ws = wb["الوصفات"]
    next_row = max(ws.max_row + 1, 3)
    num = next_row - 2
    b   = _border()
    fill = COLORS["row_odd"] if num % 2 else COLORS["row_even"]
    row_data = [
        num, r.get("title",""), r.get("meal_type",""), r.get("cuisine",""), r.get("continent",""),
        r.get("ingredients",""), r.get("steps",""), r.get("url",""),
        r.get("channel",""), r.get("duration",""), r.get("views",""),
        datetime.now().strftime("%Y-%m-%d %H:%M"), r.get("notes",""),
    ]
    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=next_row, column=col)
        cell.value = val
        cell.fill  = PatternFill("solid", start_color=fill)
        cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
        cell.border = b
        if col == 8 and val:
            cell.hyperlink = val
            cell.font = Font(size=10, color=COLORS["link"], underline="single", name="Arial")
        else:
            cell.font = Font(size=10, name="Arial")
    ws.row_dimensions[next_row].height = 90
    wb.save(EXCEL_FILE)
    return num

def get_all_recipes() -> list:
    if not os.path.exists(EXCEL_FILE):
        return []
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb["الوصفات"]
    out = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is not None:
            out.append({
                "num":row[0],"title":row[1] or "","meal_type":row[2] or "",
                "cuisine":row[3] or "","continent":row[4] or "",
                "url":row[7] or "","channel":row[8] or "",
                "duration":row[9] or "","date":row[11] or "",
            })
    return out

# ══════════════════════════════════════════════════════
#  YouTube Search
# ══════════════════════════════════════════════════════
def search_youtube(query: str, max_results: int = 6) -> list:
    try:
        import yt_dlp
        opts = {"quiet":True,"no_warnings":True,"extract_flat":True,"ignoreerrors":True}
        with yt_dlp.YoutubeDL(opts) as ydl:
            info = ydl.extract_info(f"ytsearch{max_results}:{query} recipe", download=False)
        if not info or "entries" not in info:
            raise ValueError("no entries")
        videos = []
        for e in info["entries"]:
            if not e: continue
            secs  = e.get("duration") or 0
            mins  = round(secs / 60)
            vc    = e.get("view_count") or 0
            views = f"{vc/1e6:.1f}M" if vc>=1e6 else (f"{vc/1e3:.0f}K" if vc>=1e3 else (str(vc) if vc else "N/A"))
            vid   = e.get("id","")
            url   = e.get("url") or e.get("webpage_url") or f"https://www.youtube.com/watch?v={vid}"
            videos.append({
                "title":   e.get("title",""),
                "url":     url,
                "channel": e.get("uploader") or e.get("channel",""),
                "duration": mins,
                "views":   views,
                "description": (e.get("description") or "")[:300],
            })
        return videos
    except Exception as ex:
        logger.error(f"yt-dlp: {ex}")
        try:
            hdrs = {"User-Agent":"Mozilla/5.0"}
            resp = requests.get("https://www.youtube.com/results",
                                params={"search_query": query+" recipe","hl":"ar"},
                                headers=hdrs, timeout=12)
            ids    = list(dict.fromkeys(re.findall(r'"videoId":"([^"]{11})"', resp.text)))
            titles = re.findall(r'"title":\{"runs":\[\{"text":"([^"]+)"', resp.text)
            chans  = re.findall(r'"ownerText":\{"runs":\[\{"text":"([^"]+)"', resp.text)
            return [{"title":titles[i] if i<len(titles) else f"وصفة {i+1}",
                     "url":f"https://www.youtube.com/watch?v={v}",
                     "channel":chans[i] if i<len(chans) else "يوتيب",
                     "duration":0,"views":"N/A","description":""}
                    for i,v in enumerate(ids[:max_results])]
        except Exception as ex2:
            logger.error(f"fallback: {ex2}")
            return []

# ══════════════════════════════════════════════════════
#  توليد الوصفة التفصيلية
# ══════════════════════════════════════════════════════
def generate_recipe_details(title: str, cuisine: str, meal_type: str) -> dict:
    if ANTHROPIC_KEY:
        try:
            prompt = (
                f"اكتب وصفة طبخ تفصيلية باللغة العربية للطبق: {title}\n"
                f"المطبخ: {cuisine} | النوع: {meal_type}\n\n"
                "أجب بـ JSON فقط بهذا الشكل الدقيق بدون أي نص آخر:\n"
                '{"ingredients":["2 كوب أرز","..."],"steps":["سخّن الزيت...","..."],"serving":"4 أشخاص","time":"45 دقيقة"}'
            )
            resp = requests.post(
                "https://api.anthropic.com/v1/messages",
                headers={"x-api-key":ANTHROPIC_KEY,"anthropic-version":"2023-06-01","content-type":"application/json"},
                json={"model":"claude-haiku-4-5-20251001","max_tokens":1200,
                      "messages":[{"role":"user","content":prompt}]},
                timeout=25,
            )
            raw  = resp.json()["content"][0]["text"].strip()
            raw  = re.sub(r"```json|```","",raw).strip()
            data = json.loads(raw)
            return {
                "ingredients": data.get("ingredients", []),
                "steps":       data.get("steps", []),
                "serving":     data.get("serving",""),
                "time":        data.get("time",""),
            }
        except Exception as e:
            logger.warning(f"Claude API: {e}")

    # fallback بدون API
    return {
        "ingredients":["شاهد الفيديو لقائمة المكونات الكاملة"],
        "steps":["شاهد الفيديو لطريقة التحضير خطوة بخطوة"],
        "serving":"","time":"",
    }

def format_recipe_message(video: dict, cuisine: str, meal_type: str, details: dict) -> str:
    title     = video.get("title","")
    channel   = video.get("channel","")
    url       = video.get("url","")
    duration  = video.get("duration",0)
    views     = video.get("views","N/A")
    continent = CONTINENT_MAP.get(cuisine,"")
    ingredients = details.get("ingredients",[])
    steps       = details.get("steps",[])
    serving     = details.get("serving","")
    prep_time   = details.get("time","")

    msg = f"🍽️ <b>{title}</b>\n"
    if cuisine:   msg += f"🌍 الدولة: <b>{cuisine}</b>\n"
    if continent: msg += f"🗺️ القارة: <b>{continent}</b>\n"
    if meal_type: msg += f"🍴 النوع: <b>{meal_type}</b>\n"

    info_parts = [f"📺 {channel}"]
    if duration: info_parts.append(f"⏱ {duration} دقيقة")
    if views and views != "N/A": info_parts.append(f"👁 {views}")
    msg += "\n" + "  |  ".join(info_parts) + "\n"
    if serving:   msg += f"👥 يكفي: {serving}\n"
    if prep_time: msg += f"⏳ وقت التحضير: {prep_time}\n"

    if ingredients:
        msg += "\n<b>🛒 المكونات:</b>\n"
        for ing in ingredients:
            msg += f"  • {ing}\n"

    if steps:
        msg += "\n<b>👨‍🍳 طريقة التحضير (مرحلة بمرحلة):</b>\n"
        for i, step in enumerate(steps, 1):
            msg += f"\n<b>{i}.</b> {step}\n"

    msg += f"\n🔗 <a href='{url}'>▶️ مشاهدة الفيديو على يوتيب</a>\n"
    msg += "\n💾 <b>هل تريد حفظ هذه الوصفة في Excel؟</b>"
    return msg

# ══════════════════════════════════════════════════════
#  Telegram Helpers
# ══════════════════════════════════════════════════════
def tg(method, **kw):
    try:
        return requests.post(f"{API}/{method}", json=kw, timeout=15).json()
    except Exception as e:
        logger.error(f"tg/{method}: {e}"); return {}

def send(cid, text, markup=None):
    p = {"chat_id":cid,"text":text,"parse_mode":"HTML","disable_web_page_preview":True}
    if markup: p["reply_markup"] = markup
    tg("sendMessage", **p)

def send_doc(cid, path, caption=""):
    try:
        with open(path,"rb") as f:
            requests.post(f"{API}/sendDocument",
                          data={"chat_id":cid,"caption":caption,"parse_mode":"HTML"},
                          files={"document":f}, timeout=30)
    except Exception as e:
        logger.error(f"send_doc: {e}")

def kb(rows):
    return {"keyboard":rows,"resize_keyboard":True,"one_time_keyboard":True}

def rm_kb():
    return {"remove_keyboard":True}

def main_kb():
    return kb([["🔍 بحث عن وصفة","📋 وصفاتي"],["📥 تحميل Excel","📊 إحصائيات"],["❓ مساعدة"]])

def yes_no_kb():
    return kb([["✅ نعم، احفظها","❌ لا، شكراً"]])

# ══════════════════════════════════════════════════════
#  أوامر
# ══════════════════════════════════════════════════════
def cmd_start(cid):
    user_states[cid] = {"state":"idle","data":{}}
    send(cid,
        "👨‍🍳 <b>أهلاً بك في بوت وصفات الطعام العالمي!</b>\n\n"
        "اكتب اسم أي وجبة وسأجلب لك:\n"
        "• فيديو يوتيب 🎥\n"
        "• المكونات الكاملة 🛒\n"
        "• طريقة التحضير خطوة بخطوة 👨‍🍳\n"
        "• حفظ في Excel 📊\n\n"
        "<b>مثال:</b> كبسة | بيتزا | سوشي | برياني",
        markup=main_kb())

def cmd_help(cid):
    send(cid,
        "📖 <b>طريقة الاستخدام:</b>\n\n"
        "1️⃣ اكتب اسم الوجبة\n"
        "2️⃣ اختر النوع والمطبخ\n"
        "3️⃣ اختر الفيديو المناسب\n"
        "4️⃣ احصل على الوصفة كاملة\n"
        "5️⃣ احفظها في Excel بضغطة واحدة",
        markup=main_kb())

def cmd_all(cid):
    recs = get_all_recipes()
    if not recs:
        send(cid,"📭 لا توجد وصفات محفوظة بعد!",markup=main_kb()); return
    text = f"📋 <b>وصفاتي ({len(recs)} وصفة):</b>\n\n"
    for r in recs[-20:]:
        text += (f"<b>{r['num']}. {r['title']}</b>\n"
                 f"   🌍 {r['cuisine']} | 🍴 {r['meal_type']}\n"
                 f"   🔗 <a href='{r['url']}'>مشاهدة</a> | 📅 {r['date']}\n\n")
    if len(recs)>20: text += f"<i>آخر 20 من {len(recs)}</i>"
    send(cid, text, markup=main_kb())

def cmd_excel(cid):
    get_or_create_wb()
    send(cid,"📤 جاري إرسال الملف...")
    send_doc(cid, EXCEL_FILE,
             caption=f"📊 <b>ملف وصفاتك</b>\n📝 {len(get_all_recipes())} وصفة")

def cmd_stats(cid):
    recs = get_all_recipes()
    if not recs:
        send(cid,"لا توجد إحصائيات بعد!",markup=main_kb()); return
    types,cuisines = {},{}
    for r in recs:
        t=r["meal_type"] or "غير محدد"; c=r["cuisine"] or "غير محدد"
        types[t]=types.get(t,0)+1; cuisines[c]=cuisines.get(c,0)+1
    text=f"📊 <b>إحصائياتك:</b>\n\n📝 الإجمالي: <b>{len(recs)}</b>\n\n🍴 <b>الأنواع:</b>\n"
    for t,n in sorted(types.items(),key=lambda x:-x[1])[:5]: text+=f"  • {t}: {n}\n"
    text+="\n🌍 <b>المطابخ:</b>\n"
    for c,n in sorted(cuisines.items(),key=lambda x:-x[1])[:5]: text+=f"  • {c}: {n}\n"
    send(cid,text,markup=main_kb())

# ══════════════════════════════════════════════════════
#  تدفق البحث
# ══════════════════════════════════════════════════════
def ask_meal_type(cid, name):
    rows = [MEAL_TYPES[i:i+3] for i in range(0,len(MEAL_TYPES),3)]
    rows.append(["تخطي"])
    send(cid,f"🍴 <b>نوع وجبة '{name}'؟</b>",markup=kb(rows))

def ask_cuisine(cid):
    rows = [WORLD_CUISINES[i:i+4] for i in range(0,len(WORLD_CUISINES),4)]
    rows.append(["تخطي"])
    send(cid,"🌍 <b>من أي دولة / مطبخ؟</b>",markup=kb(rows))

def do_search(cid, si):
    d       = si["data"]
    name    = d.get("name","")
    cuisine = d.get("cuisine","")
    mtype   = d.get("meal_type","")
    query   = " ".join(p for p in [name, cuisine, mtype] if p)

    send(cid,f"🔍 جاري البحث: <b>{query}</b>\n⏳ انتظر...",markup=rm_kb())
    videos = search_youtube(query)

    if not videos:
        send(cid,"❌ لم أجد نتائج. جرّب كلمات أخرى.",markup=main_kb())
        user_states[cid]={"state":"idle","data":{}}; return

    search_cache[cid] = {"videos":videos,"meal_type":mtype,"cuisine":cuisine}

    text = f"🎥 <b>نتائج: {query}</b>\n\n"
    for i, v in enumerate(videos, 1):
        dur   = f"{v['duration']} د" if v["duration"] else ""
        views = v["views"] or ""
        meta  = "  |  ".join(x for x in [dur,views] if x)
        text += (f"<b>{i}. {v['title']}</b>\n"
                 f"   📺 {v['channel']}" + (f"  |  {meta}" if meta else "") +
                 f"\n   🔗 <a href='{v['url']}'>مشاهدة</a>\n\n")
    text += "📌 <b>اكتب رقم الفيديو لرؤية الوصفة كاملة:</b>"
    nums = [[str(i) for i in range(1,len(videos)+1)],["0 إلغاء"]]
    user_states[cid]={"state":"waiting_choice","data":d}
    send(cid,text,markup=kb(nums))

def show_full_recipe(cid, video, cuisine, meal_type):
    send(cid,"⏳ <b>جاري تجهيز الوصفة التفصيلية...</b>",markup=rm_kb())
    details   = generate_recipe_details(video["title"], cuisine, meal_type)
    continent = CONTINENT_MAP.get(cuisine,"")

    recipe_cache[cid] = {
        "title":       video["title"],
        "meal_type":   meal_type,
        "cuisine":     cuisine,
        "continent":   continent,
        "ingredients": "\n".join(f"• {x}" for x in details.get("ingredients",[])),
        "steps":       "\n".join(f"{i+1}. {s}" for i,s in enumerate(details.get("steps",[]))),
        "url":         video["url"],
        "channel":     video["channel"],
        "duration":    video["duration"],
        "views":       video["views"],
        "notes":       "",
    }

    msg = format_recipe_message(video, cuisine, meal_type, details)
    user_states[cid] = {"state":"waiting_save","data":{}}
    send(cid, msg, markup=yes_no_kb())

# ══════════════════════════════════════════════════════
#  معالج التحديثات
# ══════════════════════════════════════════════════════
def handle_update(update: dict):
    if "message" not in update: return
    msg  = update["message"]
    cid  = msg["chat"]["id"]
    text = msg.get("text","").strip()
    if not text: return

    si    = user_states.get(cid,{"state":"idle","data":{}})
    state = si["state"]

    if text in ("/start",):                   cmd_start(cid);  return
    if text in ("/help","❓ مساعدة"):          cmd_help(cid);   return
    if text in ("/all","📋 وصفاتي"):          cmd_all(cid);    return
    if text in ("/excel","📥 تحميل Excel"):   cmd_excel(cid);  return
    if text in ("/stats","📊 إحصائيات"):      cmd_stats(cid);  return
    if text in ("/search","🔍 بحث عن وصفة"):
        user_states[cid]={"state":"waiting_name","data":{}}
        send(cid,"🔍 <b>اكتب اسم الوجبة:</b>",markup=rm_kb()); return

    if state == "waiting_save":
        if "نعم" in text:
            r   = recipe_cache.pop(cid,{})
            num = save_recipe(r)
            send(cid,
                f"✅ <b>تم الحفظ!</b>\n\n"
                f"📌 رقم الوصفة: <b>#{num}</b>\n"
                f"🍽️ <b>{r.get('title','')}</b>\n\n"
                "📥 اضغط <b>تحميل Excel</b> للحصول على الملف",
                markup=main_kb())
        else:
            recipe_cache.pop(cid,None)
            send(cid,"تمام! يمكنك البحث عن وصفة أخرى.",markup=main_kb())
        user_states[cid]={"state":"idle","data":{}}; return

    if state == "waiting_name":
        si["data"]["name"] = text
        user_states[cid]={"state":"waiting_meal_type","data":si["data"]}
        ask_meal_type(cid,text); return

    if state == "waiting_meal_type":
        si["data"]["meal_type"] = "" if "تخطي" in text else text
        user_states[cid]={"state":"waiting_cuisine","data":si["data"]}
        ask_cuisine(cid); return

    if state == "waiting_cuisine":
        si["data"]["cuisine"] = "" if "تخطي" in text else text
        do_search(cid,si); return

    if state == "waiting_choice":
        cache  = search_cache.get(cid,{})
        videos = cache.get("videos",[])
        try:    choice = int(re.search(r"\d+",text).group())
        except: send(cid,"❌ أدخل رقماً!"); return
        if choice == 0:
            user_states[cid]={"state":"idle","data":{}}
            send(cid,"تم الإلغاء.",markup=main_kb()); return
        if not (1 <= choice <= len(videos)):
            send(cid,f"❌ اختر من 1 إلى {len(videos)}"); return
        video   = videos[choice-1]
        cuisine = cache.get("cuisine","")
        mtype   = cache.get("meal_type","")
        search_cache.pop(cid,None)
        show_full_recipe(cid,video,cuisine,mtype); return

    # إدخال مباشر
    if len(text) > 2 and not text.startswith("/"):
        user_states[cid]={"state":"waiting_meal_type","data":{"name":text}}
        ask_meal_type(cid,text)
    else:
        cmd_start(cid)

# ══════════════════════════════════════════════════════
#  HTTP Server + Long Polling
# ══════════════════════════════════════════════════════
class HealthHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        body = b"OK"
        self.send_response(200)
        self.send_header("Content-Length",str(len(body)))
        self.end_headers()
        self.wfile.write(body)
    def log_message(self,*a): pass

def run_http():
    HTTPServer(("0.0.0.0",PORT),HealthHandler).serve_forever()

def delete_webhook():
    try:
        r = requests.get(f"{API}/deleteWebhook",params={"drop_pending_updates":True},timeout=10).json()
        logger.info("webhook deleted" if r.get("ok") else f"deleteWebhook: {r}")
    except Exception as e:
        logger.error(f"deleteWebhook: {e}")

def run_polling():
    logger.info("Starting bot...")
    delete_webhook()
    time.sleep(1)
    get_or_create_wb()
    offset = 0
    while True:
        try:
            resp = requests.get(f"{API}/getUpdates",
                                params={"offset":offset,"timeout":30},timeout=35).json()
            if not resp.get("ok"):
                logger.error(f"API: {resp}"); time.sleep(3); continue
            for upd in resp.get("result",[]):
                offset = upd["update_id"]+1
                try:    handle_update(upd)
                except Exception as e: logger.error(f"handle_update: {e}",exc_info=True)
        except requests.exceptions.Timeout: continue
        except Exception as e:
            logger.error(f"polling: {e}"); time.sleep(5)

if __name__ == "__main__":
    threading.Thread(target=run_http, daemon=True).start()
    run_polling()
